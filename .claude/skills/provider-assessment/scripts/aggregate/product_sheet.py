"""Per-product detail sheet in report.md style: Overview, unified item table, bibliography."""

from __future__ import annotations

import json
import sys
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# Support both direct execution and package import
try:
    from ..constants import VERDICT_LABEL, VERDICT_FILLS
    from .scores import compute_item_score
except ImportError:
    _scripts = Path(__file__).resolve().parent.parent
    sys.path.insert(0, str(_scripts))
    from constants import VERDICT_LABEL, VERDICT_FILLS  # type: ignore[no-redef]
    from scores import compute_item_score  # type: ignore[no-redef]


def build_product_sheet(
    ws,
    assessment: dict,
    run_dir: Path,
    checklist_items: list[dict],
    categories: list[dict],
    title: str = "",
) -> None:
    """Write a single product sheet: Overview, unified item table (color-coded), Bibliography."""
    pid = assessment["product_id"]
    vendor = assessment["vendor"]
    product_name = assessment["product_name"]
    overall_notes = assessment.get("overall_notes", "")

    item_req = {it["id"]: it["requirement"] for it in checklist_items}
    per_item = {it["item_id"]: it for it in assessment["items"]}
    item_cat: dict[str, str] = {}
    for cat in categories:
        for iid in cat["item_ids"]:
            item_cat[iid] = cat["name"]

    # Load claims for evidence text (keyed by item_id when present)
    claims_path = run_dir / "claims.jsonl"
    all_claims: list[dict] = []
    item_claims: dict[str, list[dict]] = {}
    if claims_path.exists():
        with claims_path.open(encoding="utf-8") as f:
            for line in f:
                c = json.loads(line.strip())
                all_claims.append(c)
                iid = c.get("item_id", "")
                if iid:
                    item_claims.setdefault(iid, []).append(c)

    # Load evidence (keyed by evidence_id; also indexed by item_id when present)
    evidence_path = run_dir / "evidence.jsonl"
    evidence_by_id: dict[str, dict] = {}
    evidence_by_item: dict[str, list[dict]] = {}
    if evidence_path.exists():
        with evidence_path.open(encoding="utf-8") as f:
            for line in f:
                e = json.loads(line.strip())
                eid = e.get("evidence_id", "")
                if eid:
                    evidence_by_id[eid] = e
                iid = e.get("item_id", "")
                if iid:
                    evidence_by_item.setdefault(iid, []).append(e)

    # Load sources and build source_id -> 1-based bib index
    sources_path = run_dir / "sources.jsonl"
    src_to_bib: dict[str, int] = {}
    all_sources: list[dict] = []
    if sources_path.exists():
        with sources_path.open(encoding="utf-8") as f:
            for i, line in enumerate(f, 1):
                s = json.loads(line.strip())
                src_to_bib[s["source_id"]] = i
                all_sources.append(s)

    def _claim_text(c: dict) -> str:
        return c.get("claim", "") or c.get("text", "")

    def _evidence_text(iid: str, entry: dict) -> str:
        """Build evidence text: claim + evidence quotes with bib references.

        Resolves evidence through the assessment item's ``evidence_ids`` when
        claims/evidence lack an ``item_id`` (older runs), falling back to the
        item_id-keyed grouping for runs that carry it.
        """
        ev_ids = set(entry.get("evidence_ids") or [])
        for ev in evidence_by_item.get(iid, []):
            eid = ev.get("evidence_id")
            if eid:
                ev_ids.add(eid)

        parts: list[str] = []
        seen_claims: set[str] = set()
        for c in item_claims.get(iid, []):
            parts.append(_claim_text(c))
            seen_claims.add(c.get("claim_id", ""))
        for c in all_claims:
            if c.get("claim_id") in seen_claims:
                continue
            if ev_ids.intersection(c.get("evidence_ids") or []):
                parts.append(_claim_text(c))

        seen_ev: set[str] = set()
        for eid in sorted(ev_ids):
            ev = evidence_by_id.get(eid)
            if not ev:
                continue
            seen_ev.add(eid)
            quote = ev.get("quote", "")
            locator = ev.get("locator", "")
            sid = ev.get("source_id", "")
            bib_idx = src_to_bib.get(sid, "?")
            if quote:
                entry_txt = f'"{quote}"'
                if locator:
                    entry_txt += f" ({locator})"
                entry_txt += f" [{bib_idx}]"
                parts.append(entry_txt)

        return "\n".join(parts) if parts else ""

    bold = openpyxl.styles.Font(bold=True)
    wrap_top = openpyxl.styles.Alignment(wrap_text=True, vertical="top")

    # ============================================================
    # Title
    # ============================================================
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
    ws.cell(row=1, column=1, value=title or f"Product Assessment: {vendor} - {product_name}")
    ws.cell(row=1, column=1).font = openpyxl.styles.Font(bold=True, size=14)

    row = 3

    # ============================================================
    # 1. Overview
    # ============================================================
    ws.cell(row=row, column=1, value="1. Overview").font = openpyxl.styles.Font(bold=True, size=12)
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row + 3, end_column=7)
    ws.cell(row=row, column=1, value=overall_notes).alignment = wrap_top
    row += 5

    # ============================================================
    # 2. Per-Item Verdicts (single unified table)
    # ============================================================
    ws.cell(row=row, column=1, value="2. Per-Item Verdicts").font = openpyxl.styles.Font(bold=True, size=12)
    row += 1
    tbl_headers = ["Category", "ID", "Requirement", "Verdict", "Conf", "Value", "Evidence"]
    table_start = row
    for ci, h in enumerate(tbl_headers, 1):
        ws.cell(row=row, column=ci, value=h).font = bold
    row += 1

    total_w = 0.0
    for cat in categories:
        for iid in cat["item_ids"]:
            entry = per_item.get(iid)
            req = item_req.get(iid, "")
            cat_name = item_cat.get(iid, cat["name"])
            if entry is None:
                ws.cell(row=row, column=1, value=cat_name)
                ws.cell(row=row, column=2, value=iid)
                ws.cell(row=row, column=3, value=req)
                row += 1
                continue
            item_score = compute_item_score(entry)
            total_w += item_score
            verdict_key = entry.get("verdict", "unknown")
            verdict_label = VERDICT_LABEL.get(verdict_key, "?")
            fill = VERDICT_FILLS.get(verdict_key)

            ws.cell(row=row, column=1, value=cat_name)
            ws.cell(row=row, column=2, value=iid)
            ws.cell(row=row, column=3, value=req)
            vcell = ws.cell(row=row, column=4, value=verdict_label)
            if fill:
                vcell.fill = fill
            cconf = entry.get("confidence", "")
            ws.cell(row=row, column=5, value=cconf.capitalize() if cconf else "")
            nv = entry.get("numeric_value")
            unit = entry.get("unit", "")
            ws.cell(row=row, column=6, value=f"{nv} {unit}".strip() if nv is not None else "—")
            ws.cell(row=row, column=7, value=_evidence_text(iid, entry)).alignment = wrap_top
            row += 1

    # Total score row
    ws.cell(row=row, column=1, value="").font = bold
    ws.cell(row=row, column=3, value="TOTAL SCORE").font = bold
    ws.cell(row=row, column=4, value=round(total_w, 2)).font = bold

    # Apply grid borders to the data table only
    thin = openpyxl.styles.Side(style="thin")
    grid_border = openpyxl.styles.Border(left=thin, right=thin, top=thin, bottom=thin)
    for r in range(table_start, row + 1):
        for c in range(1, 8):
            ws.cell(row=r, column=c).border = grid_border

    # Add Excel table for the data section
    tbl_ref = f"A{table_start}:G{row}"
    tbl = Table(displayName=f"Items_{pid[:20].replace('-', '_')}", ref=tbl_ref)
    tbl.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
                                        showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    ws.add_table(tbl)

    row += 2

    # ============================================================
    # 3. Bibliography (1-based index)
    # ============================================================
    ws.cell(row=row, column=1, value="3. Bibliography").font = openpyxl.styles.Font(bold=True, size=12)
    row += 1

    for idx, s in enumerate(all_sources, 1):
        title = s.get("title", "")
        url = s.get("canonical_locator", "") or s.get("raw_url", "")
        ws.cell(row=row, column=1, value=f"[{idx}]")
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        ws.cell(row=row, column=2, value=title).alignment = wrap_top
        ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=7)
        ws.cell(row=row, column=5, value=url).alignment = wrap_top
        row += 1

    bib_start = row - len(all_sources)
    bib_end = row - 1
    if bib_start <= bib_end:
        for r in range(bib_start, bib_end + 1):
            for c in range(1, 8):
                ws.cell(row=r, column=c).border = grid_border

    # Column widths
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 50
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 10
    ws.column_dimensions["G"].width = 60
