"""Aggregate all per-product assessments into a multi-sheet comparison matrix (.xlsx).

Usage:
    python .claude/skills/provider-assessment/scripts/aggregate/aggregate_matrix.py --domain bsg
        [--mode any|standard|deep] [--applicable]
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

try:
    from ..constants import (
        SKILL_ROOT, domain_paths, VERDICT_FILLS, VERDICT_LABEL,
        collect_assessments, load_yaml,
    )
    from .scores import compute_total_score, build_raw_row, build_weighted_row, build_summary_row
    from .matrix import build_transposed_matrix
    from .product_sheet import build_product_sheet
    from .styling import style_sheet, adjust_column_widths
except ImportError:
    _this = Path(__file__).resolve().parent
    _scripts = _this.parent
    sys.path.insert(0, str(_this))
    sys.path.insert(0, str(_scripts))
    from constants import (  # type: ignore[no-redef]
        SKILL_ROOT, domain_paths, VERDICT_FILLS, VERDICT_LABEL,
        collect_assessments, load_yaml,
    )
    from scores import compute_total_score, build_raw_row, build_weighted_row, build_summary_row  # type: ignore[no-redef]
    from matrix import build_transposed_matrix  # type: ignore[no-redef]
    from product_sheet import build_product_sheet  # type: ignore[no-redef]
    from styling import style_sheet, adjust_column_widths  # type: ignore[no-redef]


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--domain", required=True, choices=["bsg", "microsegmentation"])
    ap.add_argument("--runs-root", type=Path, default=None)
    ap.add_argument("--checklist", type=Path, default=None)
    ap.add_argument("--out-dir", type=Path, default=None)
    ap.add_argument("--mode", choices=["any", "standard", "deep"], default="any")
    ap.add_argument(
        "--applicable", action="store_true",
        help="Include applicable-adjusted columns (Weighted Applicable %%, Applicable %%) in the score sheets",
    )
    args = ap.parse_args()

    paths = domain_paths(args.domain)
    runs_root = args.runs_root or paths["runs"]
    checklist_path = args.checklist or paths["checklist"]
    out_dir = args.out_dir or paths["out"]

    checklist = load_yaml(checklist_path)
    item_ids = [it["id"] for it in checklist["items"]]

    # Build category list with their item IDs for score breakdown
    categories = []
    for c in checklist["categories"]:
        cat_item_ids = [it["id"] for it in checklist["items"] if it["category"] == c["id"]]
        categories.append({"id": c["id"], "name": c["name"], "item_ids": cat_item_ids})

    assessments = collect_assessments(runs_root, args.mode)
    if not assessments:
        print(f"No assessment.json found under {runs_root}", file=sys.stderr)
        return 0

    # Compute total scores for ranking, then filter to top 10
    scored_assessments = []
    for path, a in assessments:
        total_w, _, _ = compute_total_score(a["items"])
        scored_assessments.append((total_w, path, a))
    scored_assessments.sort(key=lambda x: x[0], reverse=True)
    top10 = scored_assessments[:10]

    # comparison_matrix: transposed (rows = items × categories, columns = top 10 products)
    matrix_rows, matrix_fields = build_transposed_matrix(top10, checklist["items"], categories)

    # Other tables: ALL products
    summary_rows = [build_summary_row(a, args.applicable) for _, a in assessments]

    # Column headers for simplified tables: "<category name> (max <n>)"
    cat_headers = {cat["id"]: f"{cat['name']} (max {len(cat['item_ids'])})" for cat in categories}
    total_header = f"Total (max {len(item_ids)})"
    weighted_rows = [
        build_weighted_row(a, categories, cat_headers, total_header, args.applicable)
        for _, a in assessments
    ]
    raw_rows = [build_raw_row(a, categories, cat_headers, total_header) for _, a in assessments]

    # Sort default layout by score descending:
    #   Coverage Summary & Weighted Scores -> weighted score, Raw Scores -> raw score
    summary_rows.sort(key=lambda r: (next(v for k, v in r.items() if k.startswith("Weighted Score")),
                                     next(v for k, v in r.items() if k.startswith("Raw Score"))), reverse=True)
    weighted_rows.sort(key=lambda r: r[total_header], reverse=True)
    raw_rows.sort(key=lambda r: r[total_header], reverse=True)

    summary_fields = list(summary_rows[0].keys())

    weighted_fields = ["Product ID", "Vendor", "Product Name"]
    weighted_fields.extend(cat_headers[cat["id"]] for cat in categories)
    weighted_fields.append(total_header)
    if args.applicable:
        weighted_fields.append("Weighted Applicable %")

    raw_fields = ["Product ID", "Vendor", "Product Name"]
    raw_fields.extend(cat_headers[cat["id"]] for cat in categories)
    raw_fields.append(total_header)

    # Legend: verdict meanings + scoring explanation (plain text, no table)
    legend_verdicts = [
        ("Supported", "Sản phẩm hỗ trợ đầy đủ tính năng này"),
        ("Partial", "Sản phẩm hỗ trợ một phần tính năng này"),
        ("Not Supported", "Sản phẩm không hỗ trợ tính năng này"),
        ("Unknown", "Không tìm thấy thông tin xác nhận từ tài liệu"),
        ("Not Applicable", "Tính năng không áp dụng cho kiến trúc sản phẩm"),
    ]
    legend_scoring = [
        "Cách tính điểm (scoring)",
        "[+] Weighted Score của một tiêu chí = verdict_score × confidence_weight (tối đa là 1 điểm)",
        "   [1] verdict_score (Raw Score): supported = 1.0, partial = 0.5, not_supported = 0.0, unknown = 0.0, not_applicable = 0.0",
        "   [2] confidence_weight: high = 1.0, medium = 0.75, low = 0.5",
        "[+] Điểm của một category = tổng điểm các tiêu chí trong nhóm yêu cầu",
        "[+] Absolute % = tổng weighted scores/số tiêu chí",
    ]
    legend_method = [
        "Phương pháp thu thập & đánh giá (methodology)",
        "[+] Mỗi sản phẩm được đánh giá dựa trên tài liệu công khai của hãng (datasheet, tài liệu kỹ thuật, blog) và nguồn bên thứ ba.",
        "[+] Với mỗi yêu cầu trong checklist, trích dẫn (quote) từ nguồn được ghi nhận làm bằng chứng; không có bằng chứng thì verdict là Unknown.",
        "[+] Mỗi item được gán 1 trong 5 verdict kèm mức confidence (high/medium/low) dựa trên việc thông tin có được cung cấp cả từ các bên thứ ba hay chỉ từ tài liệu của hãng.",
        "[+] Điểm được tính theo tổng có trọng số (weighted sum) như mô tả ở phần scoring ở dưới.",
        "[+] Những product có điểm tổng cao nhất được đưa vào sheet chi tiết (top 10).",
    ]

    # Product-sheet title comes from the shared template's first line
    template_path = SKILL_ROOT / "templates" / "product_report.md"
    try:
        template_title = template_path.read_text(encoding="utf-8").splitlines()[0].strip()
    except OSError:
        template_title = "# Product Assessment: {VENDOR} - {PRODUCT_NAME}"

    # -----------------------------------------------------------
    # Write xlsx
    # -----------------------------------------------------------
    out_path = out_dir / "comparison_matrix.xlsx"
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Sheet 1: Methodology (plain text, no Excel table)
    ws_method = wb.create_sheet("Methodology")
    bold14 = openpyxl.styles.Font(bold=True, size=14)
    bold = openpyxl.styles.Font(bold=True)
    ws_method.cell(row=1, column=1, value="Methodology").font = bold14
    label_to_key = {v: k for k, v in VERDICT_LABEL.items()}
    row = 2
    for line in legend_method:
        ws_method.cell(row=row, column=1, value=line).font = bold if line.startswith("Phương") else openpyxl.styles.Font()
        ws_method.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        row += 1
    row += 1
    for label, meaning in legend_verdicts:
        lcell = ws_method.cell(row=row, column=1, value=label)
        lcell.font = bold
        fill = VERDICT_FILLS.get(label_to_key.get(label, ""))
        if fill:
            lcell.fill = fill
        ws_method.cell(row=row, column=2, value=meaning)
        row += 1
    row += 1
    for line in legend_scoring:
        ws_method.cell(row=row, column=1, value=line).font = bold if line.startswith("Cách") else openpyxl.styles.Font()
        ws_method.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        row += 1

    # Sheet 2: Comparison Matrix (with verdict cell colors)
    ws_matrix = wb.create_sheet("Comparison Matrix")
    ws_matrix.append(matrix_fields)
    product_ids = matrix_fields[3:]
    for r in matrix_rows:
        ws_matrix.cell(row=ws_matrix.max_row + 1, column=1, value=r.get("category", ""))
        ws_matrix.cell(row=ws_matrix.max_row, column=2, value=r.get("item_id", ""))
        ws_matrix.cell(row=ws_matrix.max_row, column=3, value=r.get("requirement", ""))
        for ci, pid in enumerate(product_ids, 4):
            label = r.get(pid, "")
            cell = ws_matrix.cell(row=ws_matrix.max_row, column=ci, value=label)
            vkey = label_to_key.get(label, "")
            fill = VERDICT_FILLS.get(vkey)
            if fill:
                cell.fill = fill

    # Sheet 3: Coverage Summary
    ws_cov = wb.create_sheet("Coverage Summary")
    ws_cov.append(summary_fields)
    for r in summary_rows:
        ws_cov.append([r.get(f, "") for f in summary_fields])

    # Sheet 4: Raw Scores
    ws_raw = wb.create_sheet("Raw Scores")
    ws_raw.append(raw_fields)
    for r in raw_rows:
        ws_raw.append([r.get(f, "") for f in raw_fields])

    # Sheet 5: Weighted Scores
    ws_ws = wb.create_sheet("Weighted Scores")
    ws_ws.append(weighted_fields)
    for r in weighted_rows:
        ws_ws.append([r.get(f, "") for f in weighted_fields])

    # Sheets 6-15: Top 10 product detail sheets
    for _, path, a in top10:
        pid = a["product_id"]
        run_dir = path.parent
        sheet_name = pid[:31]
        ws = wb.create_sheet(sheet_name)
        title = (
            template_title
            .replace("{VENDOR}", a.get("vendor", ""))
            .replace("{PRODUCT_NAME}", a.get("product_name", ""))
        )
        build_product_sheet(ws, a, run_dir, checklist["items"], categories, title=title)

    # -----------------------------------------------------------
    # Styling: borders, alignment, tables, column widths
    # -----------------------------------------------------------
    table_names = {"Comparison Matrix", "Coverage Summary", "Raw Scores", "Weighted Scores"}
    table_style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
                                 showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    for ws in wb.worksheets:
        if ws.title in table_names:
            style_sheet(ws, product_sheet=False)
            if ws.max_row > 1:
                ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
                t = Table(displayName=ws.title.replace(" ", "_"), ref=ref)
                t.tableStyleInfo = table_style
                ws.add_table(t)
        else:
            style_sheet(ws, product_sheet=True)

    adjust_column_widths(wb)

    wb.save(out_path)

    print(f"Wrote {out_path}")
    print(f"[+] Methodology          : {len(legend_verdicts)} verdict labels + scoring & method guide")
    print(f"[+] Comparison Matrix    : {len(matrix_rows)} items × {len(top10)} products (transposed)")
    for _, _, a in top10:
        print(f"[+] {a['product_id']:<23}: report.md style ({a['vendor']})")
    print(f"[+] Coverage Summary     : {len(summary_rows)} products")
    print(f"[+] Raw Scores           : {len(raw_rows)} products, {len(categories)} categories")
    print(f"[+] Weighted Scores      : {len(weighted_rows)} products")
    return 0


if __name__ == "__main__":
    sys.exit(main())
