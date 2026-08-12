"""Worksheet styling: borders, alignment, column widths."""

from __future__ import annotations

import openpyxl
from openpyxl.utils import get_column_letter


def style_sheet(ws, product_sheet: bool = False) -> None:
    """Apply full grid borders + top alignment + wrap-text to all populated cells.

    For product sheets (product_sheet=True), only alignment is applied;
    borders are handled separately within build_product_sheet.
    """
    thin = openpyxl.styles.Side(style="thin")
    border = openpyxl.styles.Border(left=thin, right=thin, top=thin, bottom=thin)
    alignment = openpyxl.styles.Alignment(vertical="top", wrap_text=True)
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            if cell.value is not None:
                cell.alignment = alignment
                if not product_sheet:
                    cell.border = border


def adjust_column_widths(workbook) -> None:
    """Set column widths for summary sheets (product sheets handle their own)."""
    summary_names = {"Methodology", "Comparison Matrix", "Coverage Summary", "Raw Scores", "Weighted Scores"}
    for ws in workbook.worksheets:
        if ws.title not in summary_names:
            continue
        name = ws.title
        if name == "Methodology":
            ws.column_dimensions["A"].width = 22
            ws.column_dimensions["B"].width = 150
        elif name == "Comparison Matrix":
            ws.column_dimensions["A"].width = 22
            ws.column_dimensions["B"].width = 8
            ws.column_dimensions["C"].width = 55
            for c in range(4, ws.max_column + 1):
                ws.column_dimensions[get_column_letter(c)].width = 16
        elif name == "Coverage Summary":
            ws.column_dimensions["A"].width = 20
            ws.column_dimensions["B"].width = 20
            ws.column_dimensions["C"].width = 40
            for c in range(4, ws.max_column + 1):
                ws.column_dimensions[get_column_letter(c)].width = 15
        elif name in ("Raw Scores", "Weighted Scores"):
            ws.column_dimensions["A"].width = 20
            ws.column_dimensions["B"].width = 20
            ws.column_dimensions["C"].width = 40
            for c in range(4, ws.max_column + 1):
                ws.column_dimensions[get_column_letter(c)].width = 22
