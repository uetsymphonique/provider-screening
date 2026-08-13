"""Shared constants and helpers for the provider-assessment skill.

This skill only applies to this repo (unlike deep-research, which is meant to
stay portable/self-contained), so it depends on the repo's installed
`shared` package (see pyproject.toml, `pip install -e .`) for the DOMAINS
registry / domain_dir / domain_paths instead of keeping its own copy - see
guides/append-provider-groups.md.
"""

from __future__ import annotations

import json
import sys
from pathlib import Path

import openpyxl
import yaml

from shared.domains import DOMAINS, domain_dir, domain_paths

REPO_ROOT = Path(__file__).resolve().parents[4]
SKILL_ROOT = Path(__file__).resolve().parents[1]


VERDICT_LABEL = {
    "supported": "Supported",
    "partial": "Partial",
    "not_supported": "Not Supported",
    "unknown": "Unknown",
    "not_applicable": "Not Applicable",
}

VERDICT_FILLS = {
    "supported": openpyxl.styles.PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    "partial": openpyxl.styles.PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    "not_supported": openpyxl.styles.PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    "unknown": openpyxl.styles.PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
    "not_applicable": openpyxl.styles.PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"),
}


def load_yaml(path: Path):
    with path.open(encoding="utf-8") as f:
        return yaml.safe_load(f)


def load_json(path: Path):
    with path.open(encoding="utf-8") as f:
        return json.load(f)


def collect_assessments(runs_root: Path, mode_filter: str) -> list[tuple[Path, dict]]:
    if not runs_root.exists():
        return []
    results: list[tuple[Path, dict]] = []
    for path in sorted(runs_root.rglob("assessment.json")):
        try:
            data = load_json(path)
        except Exception as e:
            print(f"WARN skipping {path}: {e}", file=sys.stderr)
            continue
        if mode_filter != "any" and data.get("assessment_mode") != mode_filter:
            continue
        results.append((path, data))
    return results
